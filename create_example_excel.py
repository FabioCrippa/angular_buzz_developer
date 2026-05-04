import openpyxl
from openpyxl import Workbook

# Criar workbook
wb = Workbook()
wb.remove(wb.active)

# Turma 6A
ws1 = wb.create_sheet('6A')
ws1['A1'] = 'Nome'
ws1['B1'] = 'RA'
ws1['C1'] = 'Email'
ws1['A2'] = 'João Silva'
ws1['B2'] = 1001
ws1['C2'] = 'joao.silva@aluno.com'
ws1['A3'] = 'Maria Santos'
ws1['B3'] = 1002
ws1['C3'] = 'maria.santos@aluno.com'
ws1['A4'] = 'Pedro Costa'
ws1['B4'] = 1003
ws1['C4'] = 'pedro.costa@aluno.com'

# Turma 6B
ws2 = wb.create_sheet('6B')
ws2['A1'] = 'Nome'
ws2['B1'] = 'RA'
ws2['C1'] = 'Email'
ws2['A2'] = 'Ana Oliveira'
ws2['B2'] = 2001
ws2['C2'] = 'ana.oliveira@aluno.com'
ws2['A3'] = 'Carlos Mendes'
ws2['B3'] = 2002
ws2['C3'] = 'carlos.mendes@aluno.com'
ws2['A4'] = 'Fernanda Dias'
ws2['B4'] = 2003
ws2['C4'] = 'fernanda.dias@aluno.com'

# Turma 7A
ws3 = wb.create_sheet('7A')
ws3['A1'] = 'Nome'
ws3['B1'] = 'RA'
ws3['C1'] = 'Email'
ws3['A2'] = 'Lucas Ferreira'
ws3['B2'] = 3001
ws3['C2'] = 'lucas.ferreira@aluno.com'
ws3['A3'] = 'Beatriz Martins'
ws3['B3'] = 3002
ws3['C3'] = 'beatriz.martins@aluno.com'

# Salvar
path = r'C:\Users\cripp\projects(workspace)\angular_buzz_developer\exemplo_turmas_completo.xlsx'
wb.save(path)
print('✅ Arquivo Excel criado com sucesso!')
print('📊 Turmas: 6A (3 alunos), 6B (3 alunos), 7A (2 alunos)')
print(f'📁 Caminho: {path}')
